from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "data" / "excel-import"
LONGFORM_KEYWORDS = re.compile(
    r"\b(book|handbook|manual|monograph|thesis|dissertation|proceedings|chapter|isbn)\b|"
    r"手册|专著|学位论文|博士论文|硕士论文|会议论文集|教材",
    re.IGNORECASE,
)
DOI_RE = re.compile(r"\b10\.\d{4,9}/[^\s\"<>]+", re.IGNORECASE)
SCIENCEDIRECT_PII_RE = re.compile(r"/pii/([A-Z0-9]+)", re.IGNORECASE)


@dataclass
class ExcelPaperRecord:
    row: int
    title: str
    url: str
    doi: str
    domain: str
    publisher_hint: str
    access_group: str
    item_kind_hint: str
    tags: str
    evaluation: str
    inspiration: str
    author_info: str
    features: str
    conclusions: str
    useful_figures_data: str
    notes: str


def safe_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def clean_doi(raw: str) -> str:
    raw = unquote(raw).strip()
    raw = raw.split("?", 1)[0].split("#", 1)[0]
    raw = raw.rstrip(".,;)")
    raw = re.sub(r"^(https?://(dx\.)?doi\.org/)", "", raw, flags=re.IGNORECASE)
    return raw


def extract_doi(text: str, url: str) -> str:
    for candidate in [url, text]:
        match = DOI_RE.search(candidate or "")
        if match:
            return clean_doi(match.group(0))
    parsed = urlparse(url)
    if "doi" in parsed.path.lower():
        path = parsed.path
        marker = "/doi/"
        if marker in path:
            return clean_doi(path.split(marker, 1)[1])
    return ""


def publisher_hint(domain: str, url: str) -> str:
    domain = domain.lower()
    if "sciencedirect.com" in domain:
        return "ScienceDirect"
    if "springer.com" in domain:
        return "Springer"
    if "wiley.com" in domain:
        return "Wiley"
    if "cnki.net" in domain:
        return "CNKI"
    if "mdpi.com" in domain:
        return "MDPI"
    if "acs.org" in domain:
        return "ACS"
    if "frontiersin.org" in domain:
        return "Frontiers"
    if "clarivate" in domain or "webofscience" in domain:
        return "Web of Science"
    if not domain and url:
        return "URL"
    return domain or "unknown"


def access_group_for(domain: str, publisher: str, url: str) -> str:
    domain = domain.lower()
    if "springer.com" in domain or "mdpi.com" in domain or "frontiersin.org" in domain:
        return "auto_pdf_candidate"
    if "sciencedirect.com" in domain or "wiley.com" in domain or "cnki.net" in domain or "acs.org" in domain:
        return "manual_or_institution_required"
    if "clarivate" in domain or "webofscience" in domain:
        return "metadata_only"
    if url:
        return "needs_review"
    return "title_or_note_only"


def item_kind_hint(title: str, url: str, row_values: list[str]) -> str:
    blob = " ".join([title, url, *row_values])
    if LONGFORM_KEYWORDS.search(blob):
        return "longform_or_book_candidate"
    return "article_candidate"


def note_from_row(record: dict[str, str]) -> str:
    parts = []
    labels = [
        ("标记", "tags"),
        ("评价", "evaluation"),
        ("启发", "inspiration"),
        ("作者信息", "author_info"),
        ("特点", "features"),
        ("结论", "conclusions"),
        ("可供参考的图和数据", "useful_figures_data"),
    ]
    for label, key in labels:
        value = record.get(key, "")
        if value:
            parts.append(f"{label}: {value}")
    return "\n".join(parts)


def ris_escape(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def write_ris(records: list[ExcelPaperRecord], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for record in records:
            if not (record.title or record.url or record.doi):
                continue
            handle.write("TY  - JOUR\n")
            if record.title:
                handle.write(f"T1  - {ris_escape(record.title)}\n")
            if record.doi:
                handle.write(f"DO  - {record.doi}\n")
            if record.url:
                handle.write(f"UR  - {record.url}\n")
            if record.publisher_hint:
                handle.write(f"PB  - {record.publisher_hint}\n")
            if record.tags:
                for tag in re.split(r"[,，;/；、\s]+", record.tags):
                    tag = tag.strip()
                    if tag:
                        handle.write(f"KW  - {tag}\n")
            if record.notes:
                for line in record.notes.splitlines():
                    if line.strip():
                        handle.write(f"N1  - {ris_escape(line)}\n")
            handle.write(f"N1  - Excel row: {record.row}\n")
            handle.write("ER  - \n\n")


def read_records(excel_path: Path, sheet_name: str | None = None) -> list[ExcelPaperRecord]:
    workbook = load_workbook(excel_path, read_only=False, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    records: list[ExcelPaperRecord] = []
    for row in range(2, worksheet.max_row + 1):
        first = worksheet.cell(row, 1)
        title = safe_text(first.value)
        url = safe_text(first.hyperlink.target if first.hyperlink else "")
        row_values = [safe_text(worksheet.cell(row, col).value) for col in range(1, min(worksheet.max_column, 8) + 1)]
        if not title and not url and not any(row_values):
            continue
        parsed = urlparse(url)
        domain = parsed.netloc.lower()
        publisher = publisher_hint(domain, url)
        doi = extract_doi(title, url)
        raw = {
            "tags": safe_text(worksheet.cell(row, 2).value),
            "evaluation": safe_text(worksheet.cell(row, 3).value),
            "inspiration": safe_text(worksheet.cell(row, 4).value),
            "author_info": safe_text(worksheet.cell(row, 5).value),
            "features": safe_text(worksheet.cell(row, 6).value),
            "conclusions": safe_text(worksheet.cell(row, 7).value),
            "useful_figures_data": safe_text(worksheet.cell(row, 8).value),
        }
        records.append(
            ExcelPaperRecord(
                row=row,
                title=title,
                url=url,
                doi=doi,
                domain=domain,
                publisher_hint=publisher,
                access_group=access_group_for(domain, publisher, url),
                item_kind_hint=item_kind_hint(title, url, row_values),
                tags=raw["tags"],
                evaluation=raw["evaluation"],
                inspiration=raw["inspiration"],
                author_info=raw["author_info"],
                features=raw["features"],
                conclusions=raw["conclusions"],
                useful_figures_data=raw["useful_figures_data"],
                notes=note_from_row(raw),
            )
        )
    return records


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Prepare an old Excel paper list for Zotero/PaperQA import.")
    parser.add_argument("--excel", type=Path, required=True)
    parser.add_argument("--sheet", default="")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--format", choices=["text", "json"], default="text")
    args = parser.parse_args()

    records = read_records(args.excel, args.sheet or None)
    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    record_dicts = [asdict(record) for record in records]
    (output_dir / "excel_papers_manifest.json").write_text(
        json.dumps(record_dicts, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    fieldnames = list(record_dicts[0].keys()) if record_dicts else [field.name for field in ExcelPaperRecord.__dataclass_fields__.values()]
    write_csv(output_dir / "excel_papers_manifest.csv", record_dicts, fieldnames)
    write_csv(
        output_dir / "manual_access_required.csv",
        [asdict(record) for record in records if record.access_group == "manual_or_institution_required"],
        fieldnames,
    )
    write_csv(
        output_dir / "auto_pdf_candidates.csv",
        [asdict(record) for record in records if record.access_group == "auto_pdf_candidate"],
        fieldnames,
    )
    write_csv(
        output_dir / "book_or_longform_candidates.csv",
        [asdict(record) for record in records if record.item_kind_hint == "longform_or_book_candidate"],
        fieldnames,
    )
    doi_values = sorted({record.doi for record in records if record.doi})
    (output_dir / "doi_list.txt").write_text("\n".join(doi_values) + ("\n" if doi_values else ""), encoding="utf-8")
    urls = [record.url for record in records if record.url]
    (output_dir / "url_list.txt").write_text("\n".join(urls) + ("\n" if urls else ""), encoding="utf-8")
    write_ris(records, output_dir / "zotero_all_records_with_excel_notes.ris")
    write_ris([record for record in records if not record.doi], output_dir / "zotero_non_doi_placeholders.ris")
    readme = f"""# Excel Paper Import

Generated from:

```text
{args.excel}
```

Recommended workflow:

1. Review `summary.json` and `excel_papers_manifest.csv`.
2. In Zotero, import high-quality metadata first:
   - Use `doi_list.txt` with Zotero's Add Item by Identifier for DOI-backed records.
   - Use `zotero_non_doi_placeholders.ris` for records without DOI.
   - Do not import both `doi_list.txt` and `zotero_all_records_with_excel_notes.ris` into the same Zotero library unless you plan to merge duplicates.
   - `zotero_all_records_with_excel_notes.ris` is kept as a complete fallback/archive with Excel notes.
3. Use Zotero Connector / Find Available PDFs to attach PDFs.
4. Use `manual_access_required.csv` for records likely needing institution login, VPN, CNKI, ScienceDirect, Wiley, or ACS access.
5. Use `auto_pdf_candidates.csv` for records likely suitable for automatic public PDF checks.
6. Review `book_or_longform_candidates.csv`; these should not enter the normal paper indexing path.

Do not directly edit Zotero's SQLite database. Let Zotero import and manage items.
"""
    (output_dir / "README.md").write_text(readme, encoding="utf-8")

    summary = {
        "input": str(args.excel),
        "output_dir": str(output_dir),
        "records": len(records),
        "with_urls": sum(1 for record in records if record.url),
        "with_doi": sum(1 for record in records if record.doi),
        "access_groups": dict(Counter(record.access_group for record in records)),
        "publishers": dict(Counter(record.publisher_hint for record in records).most_common(20)),
        "longform_or_book_candidates": sum(1 for record in records if record.item_kind_hint == "longform_or_book_candidate"),
    }
    (output_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    if args.format == "json":
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    else:
        print(f"Prepared {summary['records']} records in {output_dir}")
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
